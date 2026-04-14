"""
api/index.py — O'zbek POS Tagger Backend v3.1
FastAPI  +  Groq AI  +  Rule-Based  +  Statistical Model

• Vercel serverless  →  handler = app  (avtomatik)
• Local             →  uvicorn api.index:app --host 0.0.0.0 --port 8000
"""

from __future__ import annotations

import json
import os
import re
from collections import Counter, defaultdict
from io import BytesIO
from typing import List, Optional

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

# ─── Groq ────────────────────────────────────────────────────────
try:
    from groq import Groq as _GroqClient

    _KEY = os.environ.get("GROQ_API_KEY", "")
    GROQ  = _GroqClient(api_key=_KEY)
    GROQ_MODEL = "llama-3.1-8b-instant"
    GROQ_OK = True
except Exception:
    GROQ = None
    GROQ_OK = False

# ─── openpyxl ────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

# ─────────────────────────────────────────────────────────────────
# FASTAPI APP
# ─────────────────────────────────────────────────────────────────
app = FastAPI(title="O'zbek POS Tagger", version="3.1")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─────────────────────────────────────────────────────────────────
# KONSTANTALAR
# ─────────────────────────────────────────────────────────────────
XPOS_MAP = {
    "N":"N","NER":"N","V":"V","VB":"V",
    "JJ":"ADJ","RR":"ADV","MD":"ADV",
    "Num":"NUM","NUM":"NUM","P":"P",
}
POS_LABEL = {
    "P":"Olmosh","ADV":"Ravish","ADJ":"Sifat",
    "NUM":"Son","N":"Ot","V":"Fe'l",
    "PUNCT":"Tinish","UNKNOWN":"Noma'lum",
}

PRON_KISHILIK  = {"men","sen","u","biz","siz","ular","o'zim","o'zing","o'zi","o'zimiz","o'zingiz","o'zlari"}
PRON_KORSATISH = {"bu","shu","o'sha","ushbu","mana","ana"}
PRON_SOROQ     = {"kim","nima","qayer","qaysi","qachon","qanday","necha","nechanchi"}
PRON_BELGILASH = {"hamma","barcha","hammasi","barchasi","har","har bir","har kim","har nima"}
PRON_BOLISHSIZ = {"hech kim","hech nima","hech qayer","hech qachon"}
PRON_OZLIK     = {"o'z"}
PRON_GUMON     = {"kimdir","nimadir","birov","allakim","allanima","ba'zi","ba'zilari"}

_ALL_PRONS: dict[str, str] = {}
for _p in PRON_KISHILIK:  _ALL_PRONS[_p] = "pron_kishilik"
for _p in PRON_KORSATISH: _ALL_PRONS[_p] = "pron_korsatish"
for _p in PRON_SOROQ:     _ALL_PRONS[_p] = "pron_soroq"
for _p in PRON_BELGILASH: _ALL_PRONS[_p] = "pron_belgilash"
for _p in PRON_BOLISHSIZ: _ALL_PRONS[_p] = "pron_bolishsizlik"
for _p in PRON_OZLIK:     _ALL_PRONS[_p] = "pron_ozlik"
for _p in PRON_GUMON:     _ALL_PRONS[_p] = "pron_gumon"

KNOWN_ADV = {
    "bugun","erta","kecha","hozir","doim","hali","ko'p","oz","juda","sal","ancha",
    "sekin","tez","baland","past","albatta","faqat","ham","yana","endi","birdaniga",
    "to'satdan","shunday","xuddi","nihoyatda","qattiq","tinch","ertaroq","kechroq",
    "oldin","keyin","asta","ohista","shoshib","darhol","birdan","hanuz","doimo",
}
ADJ_SUF  = ["roq","mtir","simon","dor","chan","li","siz","gi","dagi","kor","par","zor"]
NUM_WORDS = {
    "nol","bir","ikki","uch","to'rt","besh","olti","yetti","sakkiz","to'qqiz","o'n",
    "yigirma","o'ttiz","qirq","ellik","oltmish","yetmish","sakson","to'qson",
    "yuz","ming","million","milliard","birinchi","ikkinchi","uchinchi","to'rtinchi",
}
HISOB = ["ta","tasi","tadan","nafar","dona","litr","kg","km","metr","gramm","sm"]
SUFFIXES = [
    "larimizdan","larimizga","larimizni","larimizda","larining","laridan","lariga","larini","larida",
    "imizdan","imizga","imizni","imizda","imizning","ingizdan","ingizga","ingizni","ingizda",
    "lardan","larga","larni","larda","ning","dan","ga","ni","da","ka","qa",
    "miz","ngiz","lari","lar","im","ng","si","i","m","roq","mtir","gina","dir","mi","chi","oq",
]

# ─────────────────────────────────────────────────────────────────
# DATABASE YUKLASH  (lazy, bir marta)
# ─────────────────────────────────────────────────────────────────
_DB: dict   = {}
_STAT: dict = {}
_READY      = False


def _norm(text: str) -> str:
    t = str(text).lower().strip()
    t = re.sub(r"[\u2018\u2019\u02bb\u02bc'`\"]", "", t)
    t = re.sub(r"[.,!?;:()\[\]{}—\-«»]", "", t)
    return t.strip()


def _find_data_dir() -> Optional[str]:
    """Vercel, Railway, Render va lokal uchun data/ papkasini topadi."""
    here  = os.path.dirname(os.path.abspath(__file__))   # api/
    tries = [
        os.path.join(here, "..", "data"),   # ../data  (lokal va Vercel)
        os.path.join(here, "..", "..", "data"),
        os.path.join(os.getcwd(), "data"),
        "/var/task/data",                    # Vercel Lambda path
        "data",
    ]
    for p in tries:
        p = os.path.realpath(p)
        if os.path.isdir(p):
            return p
    return None


def _load_db():
    global _DB, _STAT, _READY
    if _READY:
        return

    data_dir = _find_data_dir()
    if not data_dir:
        _READY = True
        return

    suf_cnt = defaultdict(Counter)

    for fname in sorted(os.listdir(data_dir)):
        if not fname.endswith(".json"):
            continue
        path = os.path.join(data_dir, fname)
        try:
            with open(path, encoding="utf-8") as f:
                raw = json.load(f)
        except Exception:
            continue

        arr = raw if isinstance(raw, list) else next(
            (v for v in raw.values() if isinstance(v, list)), [])

        for item in arr:
            if not item or not isinstance(item, dict):
                continue
            form = str(item.get("FORM", "")).strip()
            if not form or form in ("FORM", "—"):
                continue
            xpos = str(item.get("XPOS", "")).strip()
            pos  = XPOS_MAP.get(xpos, "")
            key  = _norm(form)
            if not key:
                continue

            entry = {
                "pos":   pos or "N",
                "lemma": _norm(str(item.get("LEMMA", form))),
                "feats": str(item.get("FEATS", "∅")).strip(),
                "raw":   {
                    k: v for k, v in item.items()
                    if k not in ("FORM",) and v
                    and str(v).strip() not in ("", "—", "∅", "nan", "FEATS")
                },
            }
            _DB.setdefault(key, []).append(entry)

            if pos and len(key) >= 4:
                for n in (2, 3, 4):
                    suf_cnt[key[-n:]][pos] += 1

    # Statistik model
    for suf, cnt in suf_cnt.items():
        total = sum(cnt.values())
        if total < 5:
            continue
        best, best_c = cnt.most_common(1)[0]
        conf = best_c / total
        if conf >= 0.65:
            _STAT[suf] = {"pos": best, "conf": round(conf, 3)}

    _READY = True


def _ensure_loaded():
    if not _READY:
        _load_db()


# ─────────────────────────────────────────────────────────────────
# TAHLIL YORDAMCHI FUNKSIYALAR
# ─────────────────────────────────────────────────────────────────
def _pick(entries: list, prefer_pos: str = "") -> Optional[dict]:
    if not entries:
        return None
    if prefer_pos:
        for e in entries:
            if e["pos"] == prefer_pos:
                return e
    return max(entries, key=lambda e: (e["pos"] != "N", e.get("feats","∅") not in ("∅","")))


def _tok(token, stem, pos, subtype, conf, rule) -> dict:
    return {"token": token, "stem": stem, "pos": pos,
            "subtype": subtype, "confidence": conf, "rule": rule, "db": {}}


def _make(word, stem, entry, rule) -> dict:
    return {
        "token": word, "stem": entry.get("lemma") or stem,
        "pos": entry["pos"], "subtype": POS_LABEL.get(entry["pos"], ""),
        "confidence": 0.95, "rule": rule,
        "db": entry.get("raw", {}),
    }


def _pron_sub(rule: str) -> str:
    return {
        "pron_kishilik":     "Kishilik olmoshi",
        "pron_korsatish":    "Ko'rsatish olmoshi",
        "pron_soroq":        "So'roq olmoshi",
        "pron_belgilash":    "Belgilash olmoshi",
        "pron_bolishsizlik": "Bo'lishsizlik olmoshi",
        "pron_ozlik":        "O'zlik olmoshi",
        "pron_gumon":        "Gumon olmoshi",
    }.get(rule, "Olmosh")


# ─────────────────────────────────────────────────────────────────
# ASOSIY TAGGER
# ─────────────────────────────────────────────────────────────────
def tag_token(word: str) -> dict:
    key = _norm(word)

    if not key:
        return _tok(word, word, "PUNCT", "", 1.0, "punct")
    if re.match(r"^[.,!?;:()\[\]{}«»—\-\"']+$", key):
        return _tok(word, word, "PUNCT", "", 1.0, "punct")
    if re.match(r"^\d+([.,]\d+)*$", key):
        return _tok(word, word, "NUM", "Raqam", 1.0, "digit")
    if key in _ALL_PRONS:
        r = _ALL_PRONS[key]
        return _tok(word, key, "P", _pron_sub(r), 0.99, r)
    if key in KNOWN_ADV:
        return _tok(word, key, "ADV", "Ravish", 0.97, "adv_exact")
    if key in NUM_WORDS:
        return _tok(word, key, "NUM", "Son", 0.99, "num_exact")

    # DB — to'g'ridan-to'g'ri
    if key in _DB:
        e = _pick(_DB[key])
        if e:
            return _make(word, key, e, "database")

    # Hisob so'z (faqat aniq "50ta", "beshta")
    for hs in HISOB:
        if key.endswith(hs) and len(key) > len(hs) + 1:
            root = key[:-len(hs)]
            if root.isdigit() or root in NUM_WORDS:
                return _tok(word, root, "NUM", "Hisob so'z", 0.92, "num_hisob")

    # Sifat suffikslari
    for suf in ADJ_SUF:
        if key.endswith(suf) and len(key) > len(suf) + 1:
            sub = {"roq": "Orttirma daraja", "mtir": "Ozaytirma daraja"}.get(suf, "Sifat")
            rule_name = "adj_orttirma" if suf == "roq" else "adj_ozaytirma" if suf == "mtir" else "adj_exact"
            return _tok(word, key[:-len(suf)], "ADJ", sub, 0.82, rule_name)

    # Suffiks olib DB qidirish
    for suf in sorted(SUFFIXES, key=len, reverse=True):
        sk = _norm(suf)
        if len(key) > len(sk) + 1 and key.endswith(sk):
            root = key[:-len(sk)]
            if len(root) >= 2 and root in _DB:
                e = _pick(_DB[root])
                if e:
                    t = _make(word, root, e, f"database+{suf}")
                    t["stem"] = e.get("lemma") or root
                    return t

    # Statistik model
    for n in (4, 3, 2):
        if len(key) >= n:
            sm = _STAT.get(key[-n:])
            if sm:
                return _tok(word, key, sm["pos"],
                            POS_LABEL.get(sm["pos"], ""),
                            sm["conf"] * 0.85, "stat_model")

    return _tok(word, key, "UNKNOWN", "", 0.0, "no_rule")


# ─────────────────────────────────────────────────────────────────
# GROQ — NOMA'LUM SO'ZLARNI TEGLASH
# ─────────────────────────────────────────────────────────────────
def _groq_fill_unknowns(tokens: list) -> list:
    if not GROQ_OK or not GROQ:
        return tokens
    idxs = [i for i, t in enumerate(tokens) if t["pos"] == "UNKNOWN"]
    if not idxs:
        return tokens
    words = [tokens[i]["token"] for i in idxs]
    prompt = (
        "O'zbek tilida quyidagi so'zlarni morfologik teglang.\n"
        'Faqat JSON array qaytaring: [{"token":"...","pos":"N|V|ADJ|ADV|NUM|P","stem":"...","subtype":"...","confidence":0.0-1.0}]\n'
        f"So'zlar: {json.dumps(words, ensure_ascii=False)}"
    )
    try:
        resp = GROQ.chat.completions.create(
            model=GROQ_MODEL,
            messages=[
                {"role": "system", "content": "Sen O'zbek tili morfologiya ekspertisan. Faqat JSON qaytargin."},
                {"role": "user", "content": prompt},
            ],
            temperature=0.1,
            max_tokens=600,
        )
        raw = resp.choices[0].message.content.strip()
        m = re.search(r"\[.*\]", raw, re.DOTALL)
        if m:
            parsed = json.loads(m.group())
            for idx, res in zip(idxs, parsed):
                if isinstance(res, dict):
                    tokens[idx].update({
                        "pos":        str(res.get("pos", "UNKNOWN")),
                        "stem":       str(res.get("stem", tokens[idx]["stem"])),
                        "subtype":    str(res.get("subtype", "")),
                        "confidence": float(res.get("confidence", 0.70)),
                        "rule":       "groq_ai",
                    })
    except Exception:
        pass
    return tokens


# ─────────────────────────────────────────────────────────────────
# PYDANTIC MODELLARI
# ─────────────────────────────────────────────────────────────────
class TagReq(BaseModel):
    text: str

class AIReq(BaseModel):
    text: str
    tokens: list
    question: str

class ExportReq(BaseModel):
    tokens: list
    filename: str = "pos_natijalar"


# ─────────────────────────────────────────────────────────────────
# API ROUTES
# ─────────────────────────────────────────────────────────────────
@app.get("/health")
def health():
    _ensure_loaded()
    return {
        "status":    "ok",
        "db_words":  len(_DB),
        "stat_sufs": len(_STAT),
        "groq":      GROQ_OK,
    }


@app.post("/api/tag")
def tag(req: TagReq):
    _ensure_loaded()
    text = req.text.strip()
    if not text:
        raise HTTPException(400, "Matn bo'sh")

    words  = text.split()
    tokens = [tag_token(w) for w in words if w]
    tokens = _groq_fill_unknowns(tokens)

    return {
        "tokens": tokens,
        "stats":  dict(Counter(t["pos"] for t in tokens)),
        "total":  len(tokens),
    }


@app.post("/api/ai")
def ai_ask(req: AIReq):
    if not GROQ_OK or not GROQ:
        raise HTTPException(503, "Groq API ulangan emas.")
    summary = ", ".join(f"{t['token']}({t['pos']})" for t in req.tokens[:40])
    prompt  = (
        f'Matn: "{req.text}"\n'
        f"Tahlil: {summary}\n\n"
        f"Savol: {req.question}\n\n"
        "O'zbek tilshunoslik nuqtai nazaridan batafsil javob ber."
    )
    resp = GROQ.chat.completions.create(
        model=GROQ_MODEL,
        messages=[
            {"role": "system", "content": "Sen O'zbek tili morfologiyasi mutaxassissan."},
            {"role": "user",   "content": prompt},
        ],
        temperature=0.4,
        max_tokens=1200,
    )
    return {"answer": resp.choices[0].message.content}


@app.post("/api/export")
def export_xlsx(req: ExportReq):
    if not XLSX_OK:
        raise HTTPException(500, "openpyxl o'rnatilmagan")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "POS_Tahlil"

    headers = ["#", "Token", "O'zak", "POS", "Turkum", "Tur", "Ishonch%", "Qoida"]
    hfill = PatternFill("solid", fgColor="7C3AED")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font = hfill, hfont
        c.alignment = Alignment(horizontal="center", vertical="center")

    COLORS = {
        "P":"D1FAE5","ADV":"DBEAFE","ADJ":"FEF3C7",
        "NUM":"EDE9FE","N":"F8FAFC","V":"FCE7F3",
        "UNKNOWN":"F3F4F6","PUNCT":"F9FAFB",
    }
    for i, t in enumerate(req.tokens, 2):
        row = [
            i-1, t.get("token",""), t.get("stem",""),
            t.get("pos",""), POS_LABEL.get(t.get("pos",""),""),
            t.get("subtype",""), round((t.get("confidence",0))*100),
            t.get("rule",""),
        ]
        fc = COLORS.get(t.get("pos",""), "FFFFFF")
        fl = PatternFill("solid", fgColor=fc)
        for col, val in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fl

    for col in ws.columns:
        mx = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = re.sub(r"[^\w\-.]", "_", req.filename) + ".xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.get("/api/rules")
def rules():
    return {
        "rules": {
            "olmosh": {
                "tag": "P", "qoida": "Kishilik, ko'rsatish, so'roq, belgilash va boshqa olmoshlar",
                "kishilik": sorted(PRON_KISHILIK), "korsatish": sorted(PRON_KORSATISH),
                "soroq": sorted(PRON_SOROQ), "belgilash": sorted(PRON_BELGILASH),
            },
            "ravish":  {"tag": "ADV", "qoida": "Ma'lum ravishlar lug'ati", "misol": sorted(KNOWN_ADV)[:20]},
            "sifat":   {"tag": "ADJ", "qoida": "Sifat qo'shimchalari", "sufikslar": ADJ_SUF},
            "son":     {"tag": "NUM", "qoida": "Raqamlar va son so'zlari", "sufikslar": HISOB, "sozlar": sorted(NUM_WORDS)},
        }
    }


# ─────────────────────────────────────────────────────────────────
# VERCEL HANDLER  (shu qator muhim — Vercel shu o'zgaruvchini qidiradi)
# ─────────────────────────────────────────────────────────────────
handler = app
