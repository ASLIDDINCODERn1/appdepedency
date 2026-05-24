"""
server.py — Morphological POS tagging v3.0
Pipeline: Rule Engine → Dataset DB → Stat Model (Stanza-like) → Unknown

Ishga tushirish:
    pip install -r requirements.txt
    set GROQ_API_KEY=gsk_...   (Windows)
    python server.py
"""

import os, re, json, io, logging
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from collections import defaultdict, Counter
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"

# ─── Groq (faqat server muhiti orqali) ───
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")  # Set GROQ_API_KEY env var
groq_client  = None
if GROQ_API_KEY:
    try:
        from groq import Groq
        groq_client = Groq(api_key=GROQ_API_KEY)
        log.info("Groq AI ulandi")
    except Exception as e:
        log.warning(f"Groq ulanmadi: {e}")

POS_UZ = {
    "P":"Olmosh","RR":"Ravish","JJ":"Sifat","NUM":"Son",
    "N":"Ot","V":"Fe'l","PUNCT":"Tinish","UNKNOWN":"Noma'lum",
}

# Dataset XPOS (turli yozilishlar) -> kanonik tag
XPOS_NORM_MAP = {
    "P":"P","p":"P","PP":"P",
    "RR":"RR","R":"RR","rr":"RR",
    "MD":"RR",                      # MD ham (modal/ravish)
    "JJ":"JJ","Adj":"JJ","J":"JJ","jj":"JJ",
    "NUM":"NUM","Num":"NUM","num":"NUM",
    "N":"N","NER":"N","n":"N","N ":"N",
    "V":"V","VB":"V","v":"V","V ":"V"," V":"V",
    "PUNCT":"PUNCT",
}
# ═══════════════════════════════════════════════════════
# RULE ENGINE  —  PDF lingvistik qoidalari (TAHRIRLANGAN V3.5)
# ═══════════════════════════════════════════════════════
class UzbekRuleEngine:

    # ── Olmosh turlari ──
    KISHILIK    = frozenset({"men","sen","u","biz","siz","ular"})
    KORSATISH   = frozenset({"u","bu","shu","o'sha","ana","mana","manovi","anovi"})
    KORSATISH_B = frozenset({"mana bu","mana shu","ana bu","ana shu"})
    SOROQ       = frozenset({"kim","nima","qanday","qanaqa","qaysi","qancha",
                             "nechta","qayerda","qachon","nega","qayer","qayda"})
    BELGILASH   = frozenset({"har","hamma","barcha","bari","jami","ba'zi"})
    BELGILASH_B = frozenset({"har kim","har nima","har qaysi","har qancha","har qanaqa",
                             "har bir","har biri","har kim-kim"})
    BOLISHSIZLIK   = frozenset({"hech"})
    BOLISHSIZLIK_B = frozenset({"hech kim","hech nima","hech qaysi",
                                "hech qancha","hech qanaqa"})
    OZLIK       = frozenset({"o'z"})
    GUMON       = frozenset({"kimdir","nimadir","qayerdir","qaysidir",
                             "qanchadir","nechtadir","allakim","allanima",
                             "allaqaysi","allaqancha","allanechta","birov","kimsa"})
    GUMON_B     = frozenset({"bir kishi","bir narsa","bir nima"})

    ALL_PRON    = frozenset().union(KISHILIK, KORSATISH, SOROQ, BELGILASH, BOLISHSIZLIK, OZLIK, GUMON)
    ALL_PRON_B  = frozenset().union(KORSATISH_B, BELGILASH_B, BOLISHSIZLIK_B, GUMON_B)

    PRON_SUF = sorted([
        "larimizdan","larimizga","larimizni","larimizda","larining",
        "laridan","lariga","larini","larida",
        "imizdan","imizga","imizni","imizda","imizning",
        "ingizdan","ingizga","ingizni","ingizda",
        "ning","dan","ga","ni","da","ka","qa",
        "im","ing","i","miz","ngiz","lari","lar",
    ], key=len, reverse=True)

    # ── Ravish turlari ──
    HOLAT_R  = frozenset({"tez","sekin","asta","yayov","piyoda","ohista",
                          "shoshilinch","jim","tinch","baland","past"})
    PAYT_R   = frozenset({"bugun","ertaga","indin","kecha","hozir","erta",
                          "kechqurun","ertalab","tunda","doim","hamisha",
                          "ba'zan","goho","hali","keyin","oldin","avval","so'ng",
                          "beri","buyon","hozirgacha","darhol","zudlik bilan"})
    ORIN_R   = frozenset({"olg'a","ichkarida","tashqarida","uzoqdan","yuqorida",
                          "quyiga","nari","oldinda","olisda","yaqindan","yuqoriga",
                          "pastga","narida","berida","atrofda","tepada",
                          "tagida","ostida","ustida","orqada"})
    MIQDOR_R = frozenset({"sal","picha","xiyol","oz","ko'p","kam","ancha",
                          "xiyla","sal-pal","sira-sira","juda","g'oyat","nihoyatda",
                          "bag'oyat","behad","o'ta","biroz","haddan","ortiqcha"})
    MAQSAD_R = frozenset({"atay","atayin","ataylab","qasddan","o'rtacha",
                          "jo'rttaga","azza-bazza","noiloj","noilojlikdan",
                          "ilojsizlikdan","majburlikdan","albatta","shubhasiz"})
    ALL_ADV  = frozenset().union(HOLAT_R, PAYT_R, ORIN_R, MIQDOR_R, MAQSAD_R)
    ALL_ADV_B = frozenset({
        "bir oz","bir pas","bir lahza","bir zum","bir zumda",
        "har doim","har kuni","har gal","har zamon","har yili","har oy",
        "har safar","har vaqt","har soat","har dam",
        "bir marta","ikki marta","uch marta","necha marta",
        "tez-tez","ora-sira","o'sha zahoti","darhol shu onda",
        "o'tgan kuni","o'tgan yili","kelgan yili",
        "zudlik bilan","birga-birga","qo'l-qo'lda",
    })
    ADV_SUF  = sorted(["chasiga","larcha","layin","namo","ona",
                       "cha","lab","dek","day","lay","siz",
                       "an","in","iga","siga"], key=len, reverse=True)

    # ── Sifat turlari ──
    RANG_TUS  = frozenset({"sariq","yashil","ko'k","jigarrang","qizil","oq",
                           "qora","zangori","binafsha","pushti","kulrang",
                           "moviy","to'q","och","qo'ng'ir"})
    MAZA_TAM  = frozenset({"shirin","achchiq","nordon","tursh","sho'r",
                           "bemaza","mazali","totli"})
    HAJM      = frozenset({"katta","kichik","baland","past","uzun","qisqa",
                           "tor","keng","yupqa","qalin","ingichka","yo'g'on",
                           "ulkan","kichkina","ulug'","buyuk","kenja"})
    HID       = frozenset({"xushbo'y","muattar","badbo'y","hidli","hidsiz"})
    XUSUSIYAT = frozenset({"kamtar","odobli","quvnoq","yaxshi","yomon",
                           "chiroyli","aqlli","mehribon","jasur","kuchli",
                           "zaif","qattiq","yumshoq","issiq","sovuq",
                           "yangi","eski","toza","iflos","go'zal",
                           "aziz","sevimli","dono","g'ayratli","ishchan",
                           "shirinso'z","muloyim","qo'rqmas","botir",
                           "kuchsiz","dangasa","qiyin","oson","erkin",
                           "umumiy","joriy","teng","mas'ul","bo'sh",
                           "online","boshqa","jismoniy","maxsus","muhim"})
    ALL_ADJ   = frozenset().union(RANG_TUS, MAZA_TAM, HAJM, HID, XUSUSIYAT)
    ORTTIRMA  = frozenset({"eng","g'oyat","juda","nihoyatda","bag'oyat",
                           "behad","tim","jiqqa","lang","o'ta"})
    OZAYTIRMA = frozenset({"sal","biroz","picha","xiyla","nim","och","xiyol","ozgina"})
    ADJ_SUF   = sorted(["dagi","roq","gi","qi","li","lik",
                        "simon","dor","chan","siz","mand"], key=len, reverse=True)

    # ── Son turlari ──
    BASIC_NUM = frozenset({"bir","ikki","uch","to'rt","besh","olti","yetti",
                           "sakkiz","to'qqiz","o'n","yigirma","o'ttiz","qirq",
                           "ellik","oltmish","yetmish","sakson","to'qson",
                           "yuz","ming","million","milliard"})
    NUM_EXTRA = frozenset({"nol","sifr","yarim","chorak","nimchorak","butun"})
    ALL_NUM   = frozenset().union(BASIC_NUM, NUM_EXTRA)
    HISOB     = frozenset({"nafar","dona","bosh","juft","xil","tur","litr",
                           "kilogram","metr","sm","km","kg","gramm","tonna"})
    NUM_TYPES = {
        "tartib":    sorted(["inchi","nchi","lamchi"], key=len, reverse=True),
        "dona":      ["ta"],
        "chama":     sorted(["tacha","larcha","lab","cha"], key=len, reverse=True),
        "jamlovchi": sorted(["ovlon","ov","ala"], key=len, reverse=True),
        "taqsim":    ["tadan"],
    }
    NUM_UNITS = frozenset({"bir","ikki","uch","to'rt","besh","olti","yetti","sakkiz","to'qqiz"})
    NUM_TENS = frozenset({"o'n","yigirma","o'ttiz","qirq","ellik","oltmish","yetmish","sakson","to'qson"})
    NUM_SCALES = frozenset({"yuz","ming","million","milliard"})
    NUM_TYPE_CODES = {
        "tartib": "[[MD]]",
        "dona": "[[NUMCount]]",
        "chama": "[[NUMQ]]",
        "jamlovchi": "[[NUMCol]]",
        "taqsim": "[[NUMDiv]]",
    }

    # ── QAT'IY OT BO'LGAN SO'ZLAR (Sifatga o'tib ketishini to'suvchi filtr) ──
    HARD_NOUNS = frozenset({
        # Joy nomlari
        "andijon", "buxoro", "farg'ona", "jizzax", "xorazm", "urganch", "namangan", 
        "navoiy", "qashqadaryo", "qarshi", "qoraqalpog'iston", "nukus", "samarqand", 
        "sirdaryo", "guliston", "surxondaryo", "termiz", "toshkent", "nurafshon",
        # Millatlar
        "o'zbek", "rus", "ingliz", "qozoq", "qirg'iz", "tojik", "turkman", "qoraqalpoq", 
        "turk", "arab", "xitoy", "koreys", "yapon", "hind", "pokistonlik", "afg'on", 
        "eronlik", "fransuz", "nemis", "italyan", "ispan", "amerikalik", "kanadalik", 
        "ukrain", "belarus", "polyak", "chex", "bolgar", "yunon", "gruzin", "arman", 
        "ozarbayjon", "mo'g'ul", "vetnamlik", "misrlik", "braziliyalik", "avstraliyalik", 
        "shved", "norveg", "fin", "ruminiyalik",
        # Maxsus otlar
        "telekanal", "tayyorgarlik", "xavfsizlik", "aholi", "koridor", "festival", "orol",
        "ko'l", "cho'l", "kanal", "manzil", "sohil", "til",
        "kitobxonlik", "do'stlik", "yoshlik", "go'zallik", "boylik", "kambag'allik", 
        "ozodlik", "tenglik", "yaxshilik", "yomonlik", "insoniylik", "mehribonlik", 
        "aqllilik", "mustaqillik", "qahramonlik", "dehqonchilik", "temirchilik", 
        "o'qituvchilik", "rahbarlik", "shifokorlik", "fuqarolik", "hamkorlik"
    })
    POSSESSIVE_NOUN_ROOTS = frozenset({
        "ko'l", "cho'l", "kanal", "manzil", "festival", "sohil", "til",
    })
    NOUN_POSSESSIVE_SUF = sorted(["lari", "imiz", "ingiz", "si", "i"], key=len, reverse=True)
    NON_ADJ_OVERRIDES = {
        "orqali": ("RR", "Ravish", "vosita bildiruvchi so'z", "non_adj_orqali"),
        "tasdiq": ("N", "Ot", "Turdoosh ot", "non_adj_noun"),
        "qarshisida": ("N", "Ot", "O'rin-payt shaklidagi ot", "non_adj_noun"),
    }

    # ── Fe'l qo'shimchalari ──
    VERB_STRONG_SUF = sorted([
        "moqchiman","moqchisan","moqchimus","moqchisiz","moqchi","moqda","moq",
        "yapman","yapsan","yapmiz","yapsiz","yapti","yap",
        "yotibman","yotibsan","yotibdi","yotibmiz","yotibsiz","yotir", "yotgan","ayotgan",
        "ganman","gansan","ganmiz","gansiz","ganlar","ibman","ibsan","ibmiz","ibsiz","ibdi",
        "ajakman","ajaksan","ajakmiz","ajaksiz","ajak",
    ], key=len, reverse=True)

    VERB_PROB_SUF = sorted([
        "magansiz","magandim","magandik","magandi","magan",
        "madingiz","madilar","madim","mading","madik","madi",
        "dingiz","dilar","dik","dim","ding","di",
        "adilar","amiz","asiz","aman","asan","adi","yamiz","yasiz","yaman","yasan","ydi",
        "iyman","iysan","iymiz","iysiz","iyadi","iydi","yman","ysan","ymiz","ysiz",
        "sangiz","salar","sam","sang","sak","sa","sinlar","sin","gin","ib","gach","guncha",
    ], key=len, reverse=True)

    VERB_ROOT_HINT = frozenset({
        "bor","kel","ket","yur","tur","yot","o'tir","ol","ber","qil","et",
        "o'qi","yoz","ayt","de","bo'l","ish","ko'r","eshit","top","yo'qot",
        "ye","ich","uxla","uyg'o","chiq","kir","qayt","boshla","tugat",
        "bajar","yig'","tarqat","o'yna","yasha","kul","yig'la","o'yla",
        "ishla","qara","yugur","sotib","sot","yur","ilg'a","uch","qo'y",
        "yub","yubor","yoz","yozil","yasal","kuyla","tingla","tuz","tutib",
        "tut","o'gir","burab","bura","uchir","hayda","tag'in","bos","urin",
        "ur","yirtma","keltirib","keltir","olib","olmoq","aylan","yoriq",
    })

    NON_VERB_ROOTS = frozenset()

    def _is_verb(self, w: str):
        if not w or w in self.HARD_NOUNS:
            return None
        if w in self.ALL_ADJ or w in self.ALL_ADV or w in self.ALL_PRON \
           or w in self.ALL_NUM or w in self.HISOB or w in self.ORTTIRMA \
           or w in self.OZAYTIRMA:
            return None

        for suf in self.VERB_STRONG_SUF:
            if w.endswith(suf) and len(w) > len(suf) + 1:
                stem = w[:-len(suf)]
                if len(stem) >= 2:
                    return stem, suf, "strong"

        for suf in self.VERB_PROB_SUF:
            if w.endswith(suf) and len(w) > len(suf) + 1:
                stem = w[:-len(suf)]
                if 2 <= len(stem) <= 7:
                    if stem in self.ALL_ADJ or stem in self.ALL_ADV \
                       or stem in self.ALL_PRON or stem in self.ALL_NUM:
                        continue
                    return stem, suf, "prob"
        return None

    def norm(self, w: str) -> str:
        w = str(w).lower().strip()
        w = re.sub(r"[.,!?;:()\[\]{}—«»\"`]", "", w)
        w = re.sub(r"[\u2018\u2019\u02bb\u02bc]", "'", w)
        return w

    def _pron_stem(self, w):
        for suf in self.PRON_SUF:
            if w.endswith(suf) and len(w) > len(suf) + 1:
                stem = w[:-len(suf)]
                if stem in self.ALL_PRON:
                    return stem, suf
        return None, None

    def _num_stem(self, w):
        for ntype, sufs in self.NUM_TYPES.items():
            for suf in sufs:
                if w.endswith(suf) and len(w) > len(suf):
                    stem = w[:-len(suf)]
                    if stem in self.ALL_NUM or re.match(r"^\d+$", stem):
                        return stem, ntype
        return None, None

    def _is_simple_digit_num(self, stem):
        s = str(stem)
        if not re.fullmatch(r"\d+", s):
            return False
        n = int(s)
        return n < 10 or (n < 100 and n % 10 == 0) or n in {100, 1000, 1000000, 1000000000}

    def _is_text_num_compound(self, stems):
        if len(stems) < 2 or any(st not in self.ALL_NUM for st in stems):
            return False
        for prev, cur in zip(stems, stems[1:]):
            if prev in self.NUM_TENS and (cur in self.NUM_UNITS or cur in self.NUM_SCALES):
                continue
            if prev in self.NUM_UNITS and cur in self.NUM_SCALES:
                continue
            if prev in self.NUM_SCALES and cur in self.ALL_NUM:
                continue
            return False
        return True

    def _noun_possessive_stem(self, w):
        for suf in self.NOUN_POSSESSIVE_SUF:
            if w.endswith(suf) and len(w) > len(suf) + 1:
                stem = w[:-len(suf)]
                if stem in self.POSSESSIVE_NOUN_ROOTS or stem in self.HARD_NOUNS:
                    return stem, suf
        return None, None

    # ── Asosiy teglovchi funksiya ──
    def tag(self, word: str, prev: str = "", nxt: str = "") -> dict:
        raw = word
        w   = self.norm(word)
        pv  = self.norm(prev)
        nx  = self.norm(nxt)

        if not w:
            return self._r(raw, w, "PUNCT", "Tinish belgisi", "", 1.0, "punct")

        if w in self.NON_ADJ_OVERRIDES:
            pos, pos_uz, subtype, rule = self.NON_ADJ_OVERRIDES[w]
            cats = self._adv_categories(w, "") if pos == "RR" else {}
            return self._r(raw, w, pos, pos_uz, subtype, 1.0, rule, {"cats": cats} if cats else None)

        # ── 1. RAQAM KO'RINISHIDAGI BILIKMALAR VA MATNLAR (RegEx) ──
        # Masalan: 1, 2025, 2025-yil, 21-asr, 5-chi, 1-ta, 3-kurs
        if re.match(r'^\d+([.,/]\d+)?(-(yil|asr|sinf|maktab|qavat|kurs|kun|chorak|bosqich|fasl|ta|chi|inchi))?$', w):
            subtype = "[[NUM]]"
            if "-ta" in w: subtype = "[[NUMCount]]"
            elif "-chi" in w or "-inchi" in w: subtype = "[[MD]]"
            elif "-yil" in w: subtype = "Yil ko'rsatkichi"
            elif "-asr" in w: subtype = "Asr ko'rsatkichi"
            cats = self._num_categories(w, "", True, False)
            if subtype.startswith("[["):
                cats[" Ma'noviy xususiyatlari"] = subtype
            return self._r(raw, w, "NUM", "Son", subtype, 1.0, "digit_regex", {"cats": cats})

        # Juft raqamli sonlar (Masalan: 5-6, 10-15, 20-30)
        if re.match(r'^\d+-\d+$', w):
            cats = self._num_categories(w, "", True, False)
            return self._r(raw, w, "NUM", "Son", cats[" Ma'noviy xususiyatlari"], 1.0, "digit_juft", {"cats": cats})

        noun_stem, noun_suf = self._noun_possessive_stem(w)
        if noun_stem:
            return self._r(raw, noun_stem, "N", "Ot", "Egalik qo'shimchali ot", 0.98, "noun_possessive+" + noun_suf)

        # ── 2. QAT'IY OT FILTRI (Xavfsizlik, Andijon, O'zbek, telekanal va b.) ──
        # Agar so'z qat'iy otlar ro'yxatida bo'lsa yoki -lik bilan tugab ro'yxatda bo'lsa, uni qat'iy Ot (N) qilamiz.
        if w in self.HARD_NOUNS or w.endswith("lik") or w.endswith("chilik") or w.endswith("korlik"):
            # Istisno: agar so'z aniq sifatlar ichida bo'lmasa, uni ot qilamiz
            if w not in self.ALL_ADJ:
                subtype = "Turdoosh ot"
                if w in {"andijon", "buxoro", "farg'ona", "jizzax", "xorazm", "urganch", "namangan", "navoiy", "qashqadaryo", "qarshi", "qoraqalpog'iston", "nukus", "samarqand", "sirdaryo", "guliston", "surxondaryo", "termiz", "toshkent", "nurafshon"}:
                    subtype = "Atoqli ot (Joy nomi)"
                elif w in {"o'zbek", "rus", "ingliz", "qozoq", "qirg'iz", "tojik", "turkman", "qoraqalpoq", "turk", "arab", "xitoy", "koreys", "yapon", "hind", "fransuz", "nemis", "italyan", "ispan"}:
                    subtype = "Atoqli ot (Millat)"
                elif w.endswith("lik") or w.endswith("chilik") or w.endswith("korlik"):
                    subtype = "Mavhum ot (-lik)"
                return self._r(raw, w, "N", "Ot", subtype, 1.0, "hard_noun_rule")

        # ── 3. OLMOSH ──
        if w in self.KISHILIK:
            cats = self._pron_categories(w, "")
            return self._r(raw, w, "P", "Olmosh", "kishilik olmoshi", 1.0, "pron_kishilik", {"cats": cats})
        if w in self.KORSATISH and w not in self.KISHILIK:
            cats = self._pron_categories(w, "")
            return self._r(raw, w, "P", "Olmosh", "ko'rsatish olmoshi", 1.0, "pron_korsatish", {"cats": cats})
        if w in self.SOROQ:
            cats = self._pron_categories(w, "")
            return self._r(raw, w, "P", "Olmosh", "so'roq olmoshi", 1.0, "pron_soroq", {"cats": cats})
        if w in self.BELGILASH:
            cats = self._pron_categories(w, "")
            return self._r(raw, w, "P", "Olmosh", "belgilash olmoshi", 1.0, "pron_belgilash", {"cats": cats})
        if w in self.BOLISHSIZLIK:
            cats = self._pron_categories(w, "")
            return self._r(raw, w, "P", "Olmosh", "bo'lishsizlik olmoshi", 1.0, "pron_bolishsizlik", {"cats": cats})
        if w in self.OZLIK:
            cats = self._pron_categories(w, "")
            return self._r(raw, w, "P", "Olmosh", "o'zlik olmoshi", 1.0, "pron_ozlik", {"cats": cats})
        if w in self.GUMON:
            cats = self._pron_categories(w, "")
            return self._r(raw, w, "P", "Olmosh", "gumon olmoshi", 1.0, "pron_gumon", {"cats": cats})

        st, sf = self._pron_stem(w)
        if st:
            cats = self._pron_categories(st, sf)
            return self._r(raw, st, "P", "Olmosh", self._pron_sub(st), 0.95, "pron+" + sf, {"cats": cats})

        # ── 4. SON ──
        if w in self.ALL_NUM:
            cats = self._num_categories(w, "", False, False)
            return self._r(raw, w, "NUM", "Son", cats[" Ma'noviy xususiyatlari"], 1.0, "num_exact", {"cats": cats})
        if pv in self.ALL_NUM and w in self.HISOB:
            cats = self._num_categories(w, "dona", False, False)
            return self._r(raw, w, "NUM", "Son", cats[" Ma'noviy xususiyatlari"], 0.92, "num_hisob", {"cats": cats})
        
        # Juft matnli sonlar uchun qoida (Masalan: bir-ikki, uch-to'rt, besh-olti, o'n-o'n besh)
        if "-" in w and any(part in self.ALL_NUM for part in w.split("-")):
            cats = self._num_categories(w, "", False, False)
            return self._r(raw, w, "NUM", "Son", cats[" Ma'noviy xususiyatlari"], 1.0, "num_juft", {"cats": cats})

        stn, nt = self._num_stem(w)
        if stn:
            cats = self._num_categories(stn, nt, False, False)
            return self._r(raw, stn, "NUM", "Son", cats[" Ma'noviy xususiyatlari"], 0.95, "num+" + nt, {"cats": cats})

        # Kasr sonlar matn ko'rinishida (Masalan: uchdan bir, to'rtdan bir, beshdan ikki)
        if w.endswith("dan") and nx in self.ALL_NUM:
            return self._r(raw, w, "NUM", "Son", "Kasr son", 0.95, "num_kasr")

        # ── 5. FE'L ──
        vb = self._is_verb(w)
        if vb:
            stem, vsuf, kind = vb
            cats = self._verb_categories(stem, vsuf)
            conf = 0.95 if kind == "strong" else 0.82
            return self._r(raw, stem, "V", "Fe'l",
                           cats.get("Zamon", "") + " · " + cats.get("Mayl", ""),
                           conf, "verb+" + vsuf, {"cats": cats})

        # ── 6. SIFAT ──
        if pv in self.ORTTIRMA and w in self.ALL_ADJ:
            cats = self._adj_categories(w, pv)
            return self._r(raw, w, "JJ", "Sifat", "orttirma daraja", 1.0, "adj_orttirma", {"cats": cats})
        if pv in self.OZAYTIRMA and w in self.ALL_ADJ:
            cats = self._adj_categories(w, pv)
            return self._r(raw, w, "JJ", "Sifat", "ozaytirma daraja", 1.0, "adj_ozaytirma", {"cats": cats})
        if w in self.ALL_ADJ:
            cats = self._adj_categories(w, pv)
            return self._r(raw, w, "JJ", "Sifat", self._adj_sub(w), 1.0, "adj_exact", {"cats": cats})

        # Sifat suffikslari
        SAFE_ADJ_SUF = ("li","lik","dor","chan","siz","simon","mand","dagi","roq")
        for suf in SAFE_ADJ_SUF:
            if w.endswith(suf) and len(w) > len(suf) + 2:
                root = w[:-len(suf)]
                # Agar ildiz yoki butun so'z qat'iy otlar ichida bo'lsa sifat qilmaymiz!
                if root in self.HARD_NOUNS or w in self.HARD_NOUNS:
                    continue
                if root in self.ALL_PRON or root in self.ALL_ADV or root in self.ALL_NUM or root in self.HISOB:
                    continue
                sub = {"roq":"qiyosiy daraja","li":"xususiyat sifati",
                       "lik":"xususiyat sifati","dagi":"o'rin-payt sifati",
                       "simon":"o'xshashlik","dor":"egalik","chan":"moyillik",
                       "siz":"yo'qlik","mand":"egalik"}.get(suf, "sifat")
                cats = self._adj_categories(root, pv, suf)
                return self._r(raw, root, "JJ", "Sifat", sub, 0.80, "adj+" + suf, {"cats": cats})

        # ── 7. RAVISH ──
        if w in self.ALL_ADV:
            cats = self._adv_categories(w, "")
            return self._r(raw, w, "RR", "Ravish", self._adv_sub(w), 1.0, "adv_exact", {"cats": cats})
        
        SAFE_ADV_SUF = ("chasiga","larcha","cha","lab")
        for suf in SAFE_ADV_SUF:
            if w.endswith(suf) and len(w) > len(suf) + 3:
                root = w[:-len(suf)]
                if root in self.HARD_NOUNS or root in self.ALL_ADJ or root in self.ALL_PRON or root in self.ALL_NUM:
                    continue
                cats = self._adv_categories(root, suf)
                return self._r(raw, root, "RR", "Ravish",
                               "yasama ravish (-" + suf + ")", 0.70, "adv+" + suf, {"cats": cats})

        return self._r(raw, w, "UNKNOWN", "Noma'lum", "", 0.0, "no_rule")

    def _r(self, token, stem, pos, pos_uz, subtype, conf, rule, extra=None):
        d = {"token": token, "stem": stem, "pos": pos, "pos_uz": pos_uz,
             "subtype": subtype, "confidence": conf, "rule": rule}
        if extra:
            d.update(extra)
        return d

    def _adj_categories(self, w, pv="", suf=""):
        if w in self.RANG_TUS or w in self.MAZA_TAM or w in self.HAJM or w in self.HID or w in self.XUSUSIYAT:
            bel = "asliy sifat"
        elif suf:
            bel = "nisbiy sifat"
        else:
            bel = "asliy sifat"

        if pv in self.ORTTIRMA:         daraja = "orttirma daraja"
        elif pv in self.OZAYTIRMA:      daraja = "ozaytirma daraja"
        elif suf == "roq":              daraja = "qiyosiy daraja"
        else:                           daraja = "oddiy"

        parts = w.split("-")
        if len(parts) == 2 and parts[0] == parts[1]: tuz = "juft (takroriy)"
        elif "-" in w:                               tuz = "qo'shma"
        elif suf:                                    tuz = "yasama"
        else:                                        tuz = "sodda"

        YASAMA = {"li","lik","dor","chan","siz","simon","mand","roq","gi","dagi","qi"}
        if suf in YASAMA:        yas = "yasama sifat"
        elif w in self.ALL_ADJ:  yas = "tub sifat"
        else:                    yas = "yasama sifat"

        if w in self.RANG_TUS:          lgm = "rang-tus sifati"
        elif w in self.MAZA_TAM:        lgm = "maza-ta'm sifati"
        elif w in self.HAJM:            lgm = "hajm-o'lcham sifati"
        elif w in self.HID:             lgm = "hid bildiruvchi sifat"
        elif w in self.XUSUSIYAT:       lgm = "xususiyat sifati"
        else:                           lgm = "xususiyat sifati"

        return {
            "Belgining xususiyati": bel,
            "Daraja":               daraja,
            "Tuzilishi":            tuz,
            "Yasalishi":            yas,
            "Sifatning LMGlari":    lgm,
        }

    def _pron_categories(self, stem, suf=""):
        if stem in self.KISHILIK:        man = "Kishilik olmoshi"
        elif stem in self.KORSATISH:     man = "Ko'rsatish olmoshi"
        elif stem in self.SOROQ:         man = "So'roq olmoshi"
        elif stem in self.BELGILASH:     man = "Belgilash olmoshi"
        elif stem in self.BOLISHSIZLIK:  man = "Bo'lishsizlik olmoshi"
        elif stem in self.OZLIK:         man = "O'zlik olmoshi"
        elif stem in self.GUMON:         man = "Gumon olmoshi"
        else:                            man = "Olmosh"

        SHAXS_SON = {"men":"Birlik","sen":"Birlik","u":"Birlik",
                    "biz":"Ko'plik","siz":"Ko'plik","ular":"Ko'plik"}
        son = SHAXS_SON.get(stem, "Birlik")
        if suf.startswith("lar"): son = "Ko'plik"

        KEL = {"ning":"Qaratqich kelishik","ni":"Tushum kelishik",
               "ga":"Jo'nalish kelishik","ka":"Jo'nalish kelishik","qa":"Jo'nalish kelishik",
               "da":"O'rin-payt kelishik","dan":"Chiqish kelishik"}
        kelishik = "Bosh kelishik"
        for k in KEL:
            if suf.endswith(k):
                kelishik = KEL[k]
                break

        EGA_SUF = {"im":"I shaxs egalik","ing":"II shaxs egalik","i":"III shaxs egalik",
                   "imiz":"I shaxs egalik (ko'plik)","ingiz":"II shaxs egalik (ko'plik)",
                   "lari":"III shaxs egalik (ko'plik)"}
        egalik = "Egalik yo'q"
        for k, v in EGA_SUF.items():
            if suf == k or suf.startswith(k):
                egalik = v
                break

        tuzilish = "qo'shma" if " " in stem or "-" in stem else "sodda"
        yasalish = "affiksatsiya" if suf else "tub"

        if stem in self.KISHILIK:           vazifa = "Ot o'rnida qo'llangan"
        elif stem in self.KORSATISH:        vazifa = "Sifat o'rnida qo'llangan"
        elif stem in self.SOROQ:
            if stem in {"qancha","nechta","necha"}: vazifa = "Son o'rnida qo'llangan"
            elif stem in {"qanday","qanaqa","qaysi"}: vazifa = "Sifat o'rnida qo'llangan"
            elif stem in {"qayerda","qachon","nega","qayer","qayda"}: vazifa = "Ravish o'rnida qo'llangan"
            else:                                     vazifa = "Ot o'rnida qo'llangan"
        elif stem in self.BELGILASH:        vazifa = "Sifat o'rnida qo'llangan"
        elif stem in self.BOLISHSIZLIK:     vazifa = "Ot o'rnida qo'llangan"
        elif stem in self.OZLIK:            vazifa = "Ot o'rnida qo'llangan"
        elif stem in self.GUMON:            vazifa = "Ot o'rnida qo'llangan"
        else:                               vazifa = "Ot o'rnida qo'llangan"

        return {
            "Olmoshlarning ma'noviy guruhlari": man,
            "Tuzilishi":                        tuzilish,
            "Yasalishi":                        yasalish,
            "Kelishik":                         kelishik,
            "Son":                              son,
            "Egalik":                           egalik,
            "Olmoshlarning gapda bajaradigan vazifasiga ko'ra turlari": vazifa,
        }

    def _adv_categories(self, w, suf=""):
        if w in self.HOLAT_R:    man = "Holat ravishi"
        elif w in self.PAYT_R:   man = "Payt ravishi"
        elif w in self.ORIN_R:   man = "O'rin ravishi"
        elif w in self.MIQDOR_R: man = "Miqdor-daraja ravishi"
        elif w in self.MAQSAD_R: man = "Maqsad-sabab ravishi"
        else:                    man = "Ravish"

        parts = w.split("-")
        if " " in w:                                  tuz = "qo'shma"
        elif len(parts) == 2 and parts[0] == parts[1]: tuz = "juft (takroriy)"
        elif "-" in w:                                 tuz = "qo'shma"
        elif suf:                                     tuz = "yasama"
        else:                                          tuz = "sodda"

        YASAMA_SUF = {"cha","lab","dek","day","lay","siz","an","in","larcha","chasiga","layin"}
        if suf in YASAMA_SUF:   yas = "affiksatsiya"
        elif w in self.ALL_ADV: yas = "tub"
        else:                   yas = "yasama"

        return {
            "Ravishning ma'noviy guruhlari": man,
            "Tuzilishi":                     tuz,
            "Yasalishi":                     yas,
            "Kelishik":                      "Bosh kelishik",
            "Son":                           "Birlik",
            "Egalik":                        "Egalik yo'q",
        }

    def _num_categories(self, stem, ntype="", is_digit=False, is_compound=False):
        man = self.NUM_TYPE_CODES.get(ntype, "[[NUM]]")

        stem_text = str(stem)
        if is_compound or " " in stem_text:
            tuz = "[[NUMCmp]]"
        elif is_digit and not self._is_simple_digit_num(stem_text):
            tuz = "[[NUMCmp]]"
        else:
            tuz = "[[NUMD]]"

        hisob = "raqam ko'rinishida" if is_digit else "—"
        bir_mano = "[[NUMC]]" if self.norm(stem_text) == "bir" else "—"

        return {
            " Ma'noviy xususiyatlari": man,
            "Hisob so'zlar":           hisob,
            "Bir so'zining ma'nolari": bir_mano,
            "Tuzilishiga ko'ra":       tuz,
        }

    def _verb_categories(self, stem, suf=""):
        PAST   = {"di","dim","ding","dik","dilar","dingiz",
                  "madi","madim","mading","madik","madilar","madingiz",
                  "magan","magandim","magandi","magandik","magansiz",
                  "ganman","gansan","ganmiz","gansiz","ganlar",
                  "ibman","ibsan","ibmiz","ibsiz","ibdi"}
        PRES   = {"yap","yapti","yapman","yapsan","yapmiz","yapsiz",
                  "yotir","yotibdi","yotibman","yotibsan","yotibmiz","yotibsiz",
                  "yotgan","ayotgan","moqda"}
        FUT    = {"ajak","ajakman","ajaksan","ajakmiz","ajaksiz",
                  "moqchi","moqchiman","moqchisan","moqchimus","moqchisiz"}
        AORIST = {"adi","aman","asan","amiz","asiz","adilar",
                  "ydi","yaman","yasan","yamiz","yasiz"}
        if suf in PAST:        zamon = "o'tgan zamon"
        elif suf in PRES:      zamon = "hozirgi zamon"
        elif suf in FUT:       zamon = "kelasi zamon"
        elif suf in AORIST:    zamon = "hozirgi-kelasi (aorist)"
        elif suf == "moq":     zamon = "noaniq (infinitiv)"
        else:                  zamon = "—"

        ORDER = {"gin","sin","sinlar"}
        COND  = {"sa","sam","sang","sak","sangiz","salar"}
        if suf in ORDER:       mayl = "buyruq-istak mayli"
        elif suf in COND:      mayl = "shart mayli"
        elif suf == "moqchi" or suf.startswith("moqchi"): mayl = "maqsad mayli"
        elif suf == "moq":     mayl = "harakat nomi"
        else:                  mayl = "xabar (aniqlik) mayli"

        S1B = {"man","yapman","yotibman","ganman","ibman","ajakman","aman","yaman","moqchiman","sam","dim","madim","gandim","magandim"}
        S2B = {"san","yapsan","yotibsan","gansan","ibsan","ajaksan","asan","yasan","moqchisan","sang","ding","mading","magansiz"}
        S3B = {"di","madi","magan","adi","ydi","yapti","yotibdi","sa","sin","gan","ib","yotgan"}
        S1K = {"miz","yapmiz","yotibmiz","ganmiz","ibmiz","ajakmiz","amiz","yamiz","moqchimiz","sak","dik","madik","gandik"}
        S2K = {"ngiz","yapsiz","yotibsiz","gansiz","ibsiz","ajaksiz","asiz","yasiz","moqchisiz","sangiz","dingiz","madingiz"}
        S3K = {"lar","dilar","madilar","adilar","ganlar","sinlar","salar"}
        
        if suf in S1B:     shaxs = "I shaxs birlik"
        elif suf in S2B:   shaxs = "II shaxs birlik"
        elif suf in S1K:   shaxs = "I shaxs ko'plik"
        elif suf in S2K:   shaxs = "II shaxs ko'plik"
        elif suf in S3K:   shaxs = "III shaxs ko'plik"
        elif suf in S3B:   shaxs = "III shaxs birlik"
        else:              shaxs = "—"

        bolishli = "bo'lishsiz (-ma-)" if suf.startswith("ma") or suf.startswith("magan") or suf.startswith("madi") else "bo'lishli"
        YASAMA_FE_SUF = {"la","lash","lan","lat","tir","dir","gaz","giz"}
        yas = "asl (sodda) fe'l"
        for ys in YASAMA_FE_SUF:
            if stem.endswith(ys) and len(stem) > len(ys) + 1:
                yas = "yasama fe'l (-" + ys + ")"
                break
        return {
            "Zamon": zamon,
            "Mayl": mayl,
            "Shaxs-son": shaxs,
            "Bo'lishli": bolishli,
            "Yasalishi": yas,
        }

    def _pron_sub(self, st):
        if st in self.KISHILIK: return "kishilik olmoshi"
        if st in self.KORSATISH: return "ko'rsatish olmoshi"
        if st in self.SOROQ: return "so'roq olmoshi"
        if st in self.BELGILASH: return "belgilash olmoshi"
        if st in self.BOLISHSIZLIK: return "bo'lishsizlik olmoshi"
        if st in self.OZLIK: return "o'zlik olmoshi"
        if st in self.GUMON: return "gumon olmoshi"
        return "olmosh"

    def _adv_sub(self, w):
        if w in self.HOLAT_R: return "holat ravishi"
        if w in self.PAYT_R: return "payt ravishi"
        if w in self.ORIN_R: return "o'rin ravishi"
        if w in self.MIQDOR_R: return "miqdor-daraja ravishi"
        if w in self.MAQSAD_R: return "maqsad-sabab ravishi"
        return "ravish"

    def _adj_sub(self, w):
        if w in self.RANG_TUS: return "rang-tus sifati"
        if w in self.MAZA_TAM: return "maza-ta'm sifati"
        if w in self.HAJM: return "hajm sifati"
        if w in self.HID: return "hid bildiruvchi sifat"
        if w in self.XUSUSIYAT: return "xususiyat sifati"
        return "sifat"
# ═══════════════════════════════════════════════════════
# STATISTICAL MODEL  —  Stanza kabi dataset'dan o'rganish
# ═══════════════════════════════════════════════════════
class DatasetStatModel:
    """
    Dataset'dagi barcha so'zlardan suffix/prefix statistikasi o'rganadi.
    Rule va DB topmaganida tahminiy POS qaytaradi.
    """
    XPOS_NORM = {
        "P":  "P",   "RR": "RR",  "MD": "RR",
        "JJ": "JJ",  "Adj":"JJ",  "Num":"NUM", "NUM":"NUM",
        "N":  "N",   "NER":"N",   "V":  "V",  "VB": "V",
    }

    def __init__(self):
        self.suf_cnt: Dict[str, Counter] = {}
        self.pfx_cnt: Dict[str, Counter] = {}
        self.trained   = False
        self._n_samples = 0

    def train(self, db: Dict[str, list]):
        suf: Dict[str, Counter] = defaultdict(Counter)
        pfx: Dict[str, Counter] = defaultdict(Counter)

        for word, entries in db.items():
            for e in entries:
                norm = self.XPOS_NORM.get(str(e.get("XPOS", "")))
                if not norm:
                    continue
                # Suffix n-gram (1-6 harf) — uzun = og'irroq
                for n in range(1, min(7, len(word) + 1)):
                    suf[word[-n:]][norm] += n * n
                # Prefix n-gram (1-4 harf)
                for n in range(1, min(5, len(word) + 1)):
                    pfx[word[:n]][norm] += n
                self._n_samples += 1

        self.suf_cnt = dict(suf)
        self.pfx_cnt = dict(pfx)
        self.trained  = True
        log.info(f"StatModel: {len(self.suf_cnt)} sufiks naqshi | {self._n_samples} namuna")

    def predict(self, word: str) -> Tuple[str, float]:
        if not self.trained or not word:
            return "N", 0.22

        w = word.lower()
        scores: Counter = Counter()

        for n in range(1, min(7, len(w) + 1)):
            suf = w[-n:]
            if suf in self.suf_cnt:
                wt = n * n
                for pos, cnt in self.suf_cnt[suf].items():
                    scores[pos] += cnt * wt

        for n in range(1, min(5, len(w) + 1)):
            pfx = w[:n]
            if pfx in self.pfx_cnt:
                for pos, cnt in self.pfx_cnt[pfx].items():
                    scores[pos] += cnt

        if not scores:
            return "N", 0.20

        total    = sum(scores.values())
        best_pos, best_cnt = scores.most_common(1)[0]
        conf     = min(0.78, best_cnt / total)
        return best_pos, round(conf, 3)


# ═══════════════════════════════════════════════════════
# DATABASE LOOKUP  —  4 JSON fayl
# ═══════════════════════════════════════════════════════
class DatabaseLookup:
    XPOS_MAP = {
        "P":  ("P",   "Olmosh"),
        "RR": ("RR",  "Ravish"), "MD": ("RR",  "Ravish"), "R": ("RR", "Ravish"),
        "JJ": ("JJ",  "Sifat"),  "Adj":("JJ",  "Sifat"),  "J": ("JJ", "Sifat"),
        "Num":("NUM", "Son"),    "NUM":("NUM", "Son"),
        "N":  ("N",   "Ot"),     "NER":("N",   "Ot"),
        "V":  ("V",   "Fe'l"),   "VB": ("V",   "Fe'l"),
    }
    SUFFIXES = sorted([
        "larimizdan","larimizga","larimizni","larimizda","larining",
        "laridan","lariga","larini","larida",
        "imizdan","imizga","imizni","imizda","imizning",
        "ingizdan","ingizga","ingizni","ingizda",
        "ning","dan","ga","ni","da","ka","qa",
        "miz","ngiz","lari","lar","im","ng","si","i","m",
        "roq","mtir","gina","dir","mi","chi",
    ], key=len, reverse=True)

    def __init__(self):
        self.db:   Dict[str, list]   = {}
        self.stat: DatasetStatModel  = DatasetStatModel()
        self._load()
        self.stat.train(self.db)

    def _norm(self, w):
        w = str(w).lower().strip()
        w = re.sub(r"[.,!?;:()\[\]{}—«»\"`]", "", w)
        w = re.sub(r"[\u2018\u2019\u02bb\u02bc]", "'", w)
        return w.strip()

    def _load(self):
        sources = [
            ("Olmoshvaravish.json", "OlmoshvaRavish"),
            ("Sifat.json",          "Sifat"),
            ("Son.json",            "Son"),
            ("database.json",       None),
        ]
        total = 0
        for fname, key in sources:
            path = DATA_DIR / fname
            if not path.exists():
                log.warning(f"Topilmadi: {fname}")
                continue
            try:
                with open(path, encoding="utf-8") as f:
                    raw = json.load(f)
                arr = raw if isinstance(raw, list) else (
                    raw.get(key) if key and key in raw else
                    next((v for v in raw.values() if isinstance(v, list)), [])
                )
                cnt = 0
                for item in arr:
                    if item is None:
                        continue
                    form = str(item.get("FORM", "")).strip()
                    if not form or form in ("FORM", "—", ""):
                        continue
                    xpos = str(item.get("XPOS", "")).strip()
                    if xpos in ("", "?", "XPOS", "C", "CONJ", "PUNCT", "II", "Prt", "UH", "U", "IM", "IB"):
                        continue
                    # Kanonik tagga aylantirish (turli yozilish: p→P, J→JJ, N →N, V →V, Adj→JJ ...)
                    norm_x = XPOS_NORM_MAP.get(xpos)
                    if norm_x:
                        item["XPOS"] = norm_x
                    else:
                        continue
                    k = self._norm(form)
                    if k:
                        self.db.setdefault(k, []).append(item)
                        cnt += 1
                total += cnt
                log.info(f"  {fname}: {cnt} yozuv")
            except Exception as e:
                log.warning(f"  {fname} xato: {e}")
        log.info(f"Jami DB: {len(self.db)} noyob so'z | {total} yozuv")

    def lookup(self, word: str) -> Optional[dict]:
        k = self._norm(word)
        if k in self.db:
            return self._best(self.db[k])
        for suf in self.SUFFIXES:
            if len(k) > len(suf) + 2 and k.endswith(suf):
                root = k[:-len(suf)]
                if root in self.db:
                    entry = dict(self._best(self.db[root]))
                    entry["_suffix"] = suf
                    return entry
        return None

    def _best(self, entries):
        def score(e):
            s = 0
            if e.get("XPOS", "") in self.XPOS_MAP: s += 4
            if e.get("FEATS", "") not in ("", "∅", "—"): s += 2
            if e.get("LEMMA", "") not in ("", "∅", "—"): s += 1
            return s
        return max(entries, key=score)

    def subtype(self, entry: dict) -> str:
        xpos = entry.get("XPOS", "")
        if xpos == "P":
            v = entry.get("Olmoshlarning ma’noviy guruhlari", "")
            return v if v and v != "—" else "olmosh"
        if xpos == "JJ":
            v = entry.get("Belgining xususiyati", entry.get("Daraja", ""))
            return v if v and v != "—" else "sifat"
        if xpos == "NUM":
            v = entry.get(" Ma'noviy xususiyatlari", entry.get("Ma’noviy xususiyatlari", ""))
            return v if v and v != "—" else "son"
        if xpos == "RR":
            return "ravish"
        return ""

    def extra_cols(self, entry: dict) -> dict:
        skip = {"FORM", "LEMMA", "FEATS", "XPOS", "ID", "_suffix"}
        ADV_COL_MAP = {
            "Column13": "Ravishning ma'noviy guruhlari",
            "Column14": "Tuzilishi",
            "Column15": "Yasalishi",
            "Column16": "Kelishik",
            "Column17": "Son",
            "Column18": "Egalik",
            "Tuzalishiga ko'ra": "Tuzilishiga ko'ra",
        }
        out = {}
        for k, v in entry.items():
            if k in skip or k.startswith("_"):
                continue
            if v in ("—", "∅", "", None):
                continue
            sv = str(v).strip()
            if sv in ("—", "∅", ""):
                continue
            nk = ADV_COL_MAP.get(k, k)
            out[nk] = sv
        return out


# ═══════════════════════════════════════════════════════
# POS TAGGER  —  4 bosqichli pipeline
# ═══════════════════════════════════════════════════════
class UzbekPOSTagger:
    def __init__(self, engine: UzbekRuleEngine, db: DatabaseLookup):
        self.e    = engine
        self.db   = db
        self.stat = db.stat

    def tokenize(self, text: str) -> List[str]:
        tokens = re.findall(
            r"\d+(?:[.,/]\d+)?(?:-\d+(?:[.,/]\d+)?)?(?:-(?:yil|asr|sinf|maktab|qavat|kurs|kun|chorak|bosqich|fasl|ta|chi|inchi))?|[\w'\u02bb\u02bc\u2018\u2019\-]+|[.,!?;:\u2014\u00ab\u00bb]",
            text, re.UNICODE
        )
        return [t.strip() for t in tokens if t.strip()]

    def tag_sentence(self, text: str) -> List[dict]:
        tokens  = self.tokenize(text)
        results = []
        i = 0

        while i < len(tokens):

            # ── 1. Birikma (2-3 token): Olmosh, Ravish ──
            hit = False
            for ln in [3, 2]:
                if i + ln <= len(tokens):
                    phrase = " ".join(tokens[i:i + ln])
                    pn     = self.e.norm(phrase)
                    # Qo'shma olmosh
                    if pn in self.e.ALL_PRON_B:
                        if pn in self.e.KORSATISH_B:    sub, mn = "Ko'rsatish olmoshi", "Ko'rsatish olmoshi"
                        elif pn in self.e.BELGILASH_B:  sub, mn = "Belgilash olmoshi",  "Belgilash olmoshi"
                        elif pn in self.e.BOLISHSIZLIK_B: sub, mn = "Bo'lishsizlik olmoshi", "Bo'lishsizlik olmoshi"
                        else:                            sub, mn = "Gumon olmoshi", "Gumon olmoshi"
                        pcats = {
                            "Olmoshlarning ma'noviy guruhlari": mn,
                            "Tuzilishi":  "qo'shma",
                            "Yasalishi":  "tub",
                            "Kelishik":   "Bosh kelishik",
                            "Son":        "Birlik",
                            "Egalik":     "Egalik yo'q",
                            "Olmoshlarning gapda bajaradigan vazifasiga ko'ra turlari": "Ot o'rnida qo'llangan",
                        }
                        results.append({
                            "token": phrase, "stem": pn,
                            "pos": "P", "pos_uz": "Olmosh",
                            "subtype": sub, "confidence": 1.0,
                            "rule": "birikma_p", "index": i, "cats": pcats,
                        })
                        i += ln; hit = True; break
                    # Qo'shma ravish
                    if pn in self.e.ALL_ADV_B:
                        # Birikmaga ko'ra ma'noviy guruhini topamiz
                        if any(x in pn for x in ("doim","kuni","gal","zamon","yili","oy","safar","vaqt","soat","dam","marta","zumda","zum","lahza","pas","kuni","zahoti","onda")):
                            mn = "Payt ravishi"
                        elif any(x in pn for x in ("oz","pas","biroz","bilan")):
                            mn = "Miqdor-daraja ravishi"
                        elif any(x in pn for x in ("birga","qo'l-qo'lda")):
                            mn = "Holat ravishi"
                        else:
                            mn = "Ravish"
                        acats = {
                            "Ravishning ma'noviy guruhlari": mn,
                            "Tuzilishi": "qo'shma",
                            "Yasalishi": "qo'shma yasalishi",
                            "Kelishik":  "Bosh kelishik",
                            "Son":       "Birlik",
                            "Egalik":    "Egalik yo'q",
                        }
                        results.append({
                            "token": phrase, "stem": pn,
                            "pos": "RR", "pos_uz": "Ravish",
                            "subtype": mn,
                            "confidence": 1.0,
                            "rule": "birikma_adv", "index": i, "cats": acats,
                        })
                        i += ln; hit = True; break
            if hit:
                continue

            # ── 1b. Qo'shma son (qo'shni sonlar): "yigirma besh", "bir yuz ellik" ──
            if i + 1 < len(tokens):
                w0 = self.e.norm(tokens[i])
                if w0 in self.e.ALL_NUM:
                    j = i + 1
                    parts = [tokens[i]]
                    while j < len(tokens):
                        wj = self.e.norm(tokens[j])
                        if wj in self.e.ALL_NUM:
                            parts.append(tokens[j]); j += 1
                        else:
                            break
                    stems = [self.e.norm(p) for p in parts]
                    if self.e._is_text_num_compound(stems):
                        phrase = " ".join(parts)
                        pn     = self.e.norm(phrase)
                        ncats  = self.e._num_categories(pn, "", False, True)
                        results.append({
                            "token": phrase, "stem": pn,
                            "pos": "NUM", "pos_uz": "Son",
                            "subtype": ncats[" Ma'noviy xususiyatlari"],
                            "confidence": 1.0,
                            "rule": "birikma_num", "index": i, "cats": ncats,
                        })
                        i = j
                        continue

            tok  = tokens[i]
            prev = tokens[i - 1] if i > 0 else ""
            nxt  = tokens[i + 1] if i + 1 < len(tokens) else ""

            # ── 2. Rule Engine (PDF qoidalari) ──
            res = self.e.tag(tok, prev, nxt)
            if res["pos"] != "UNKNOWN":
                res["index"] = i
                results.append(res)
                i += 1
                continue

            # ── 3. Database lookup ──
            entry = self.db.lookup(tok)
            if entry:
                xpos       = entry.get("XPOS", "")
                pos, pos_uz = self.db.XPOS_MAP.get(xpos, ("N", "Ot"))
                suf        = entry.get("_suffix", "")
                lemma      = self.e.norm(entry.get("LEMMA", tok))
                db_extra = self.db.extra_cols(entry)
                cats = {}
                if pos == "JJ" and not db_extra:
                    cats = self.e._adj_categories(lemma or self.e.norm(tok))
                elif pos == "RR" and not db_extra:
                    cats = self.e._adv_categories(lemma or self.e.norm(tok), "")
                results.append({
                    "token":    tok,
                    "stem":     lemma or self.e.norm(tok),
                    "pos":      pos,
                    "pos_uz":   pos_uz,
                    "subtype":  self.db.subtype(entry),
                    "confidence": 0.90 if not suf else 0.83,
                    "rule":     "database" + ("+" + suf if suf else ""),
                    "index":    i,
                    "db":       db_extra,
                    **({"cats": cats} if cats else {}),
                })
                i += 1
                continue

            # ── 4. Statistical Model (Stanza-like) ──
            stat_pos, stat_conf = self.stat.predict(tok)
            results.append({
                "token":      tok,
                "stem":       self.e.norm(tok),
                "pos":        stat_pos,
                "pos_uz":     POS_UZ.get(stat_pos, "Noma'lum"),
                "subtype":    "tahminiy (stat)",
                "confidence": stat_conf,
                "rule":       "stat_model",
                "index":      i,
            })
            i += 1

        # ── POST-PASS: qo'shma so'zlarni birlashtirish ──
        return self._merge_compounds(results)

    def _merge_compounds(self, tokens: list) -> list:
        """Qo'shni bir xil POS (NUM/RR/JJ/P) tokenlarini qo'shma so'z sifatida birlashtiradi."""
        if not tokens:
            return tokens

        out = []
        i = 0
        N = len(tokens)
        # NUM uchun aniq birikma hosil qiluvchi o'zaklar
        NUM_STEM_SET = self.e.ALL_NUM
        HISOB_SET    = self.e.HISOB
        # Qo'shma ravishda ikkinchi bo'la oladigan so'zlar (N → RR birikma)
        # Masalan: "bir oz", "bir zum", "har kuni", "o'tgan kuni" — birinchisi olmosh/sifat/son bo'lishi mumkin.

        while i < N:
            t = tokens[i]
            pos = t.get("pos", "")

            # 1) NUM birikma: faqat haqiqiy matnli qo'shma sonlar
            if pos == "NUM" and i + 1 < N and tokens[i+1].get("pos") == "NUM":
                j = i + 1
                while j < N and tokens[j].get("pos") == "NUM":
                    j += 1
                parts = tokens[i:j]
                stems = [p.get("stem", p["token"]) for p in parts]
                if not self.e._is_text_num_compound(stems):
                    out.append(t)
                    i += 1
                    continue
                phrase  = " ".join(p["token"] for p in parts)
                stem    = " ".join(stems)
                ncats = self.e._num_categories(stem, "", False, True)
                # agar ketma-ket oxirida hisob so'z bo'lsa — "dona son" sifatida belgilansin
                last = parts[-1].get("stem", "")
                if last in HISOB_SET:
                    ncats[" Ma'noviy xususiyatlari"] = self.e.NUM_TYPE_CODES["dona"]
                    ncats["Hisob so'zlar"] = last
                out.append({
                    "token": phrase, "stem": stem,
                    "pos": "NUM", "pos_uz": "Son",
                    "subtype": ncats[" Ma'noviy xususiyatlari"],
                    "confidence": 1.0,
                    "rule": "birikma_num", "index": t.get("index", i),
                    "cats": ncats,
                })
                i = j
                continue

            # 2) Tire bilan bog'langan JJ/RR: "qizil-sariq", "katta-kichik", "asta-sekin"
            #    Tokenlar allaqachon tire bilan kelgan (tokenizer ularni bir token qilgan) — bu holatga qaramaymiz.

            # 3) Juft (takroriy) JJ/RR: "katta katta", "sekin sekin"
            if pos in ("JJ", "RR") and i + 1 < N \
               and tokens[i+1].get("pos") == pos \
               and self.e.norm(t.get("token","")) == self.e.norm(tokens[i+1].get("token","")):
                phrase = t["token"] + " " + tokens[i+1]["token"]
                stem   = t.get("stem", "")
                cats = dict(t.get("cats", {}))
                cats["Tuzilishi"] = "juft (takroriy)"
                out.append({
                    "token": phrase, "stem": stem,
                    "pos": pos, "pos_uz": t.get("pos_uz", ""),
                    "subtype": t.get("subtype", ""),
                    "confidence": 1.0,
                    "rule": "birikma_juft", "index": t.get("index", i),
                    "cats": cats,
                })
                i += 2
                continue

            # 4) "har + N" → qo'shma ravish (har kuni, har gal, har safar ...)
            #    lekin "har bir", "har kim" va hokazolar — belgilash olmoshi (allaqachon birikma_p bilan ushlangan).
            if self.e.norm(t.get("token", "")) == "har" and i + 1 < N:
                nxt = tokens[i+1]
                HAR_ADV_NEXT = {"kuni","gal","zamon","yili","oy","safar","vaqt","soat","dam","payt","dafa","doim"}
                nxt_norm = self.e.norm(nxt.get("token", ""))
                if nxt_norm in HAR_ADV_NEXT:
                    phrase = t["token"] + " " + nxt["token"]
                    pn = self.e.norm(phrase)
                    acats = {
                        "Ravishning ma'noviy guruhlari": "Payt ravishi",
                        "Tuzilishi": "qo'shma",
                        "Yasalishi": "qo'shma yasalishi",
                        "Kelishik":  "Bosh kelishik",
                        "Son":       "Birlik",
                        "Egalik":    "Egalik yo'q",
                    }
                    out.append({
                        "token": phrase, "stem": pn,
                        "pos": "RR", "pos_uz": "Ravish",
                        "subtype": "Payt ravishi",
                        "confidence": 0.92,
                        "rule": "birikma_har", "index": t.get("index", i),
                        "cats": acats,
                    })
                    i += 2
                    continue

            # 5) "bir + N/NUM" → ba'zi birikmalar (bir oz, bir pas, bir zum, bir marta, bir kuni)
            if self.e.norm(t.get("token", "")) == "bir" and i + 1 < N:
                nxt_raw = self.e.norm(tokens[i+1].get("token", ""))
                BIR_ADV = {"oz","pas","zum","lahza","zumda","kuni","marta","pasda","lahzada"}
                if nxt_raw in BIR_ADV:
                    phrase = t["token"] + " " + tokens[i+1]["token"]
                    pn = self.e.norm(phrase)
                    mn = "Payt ravishi" if nxt_raw in {"kuni","marta","lahza","lahzada","zum","zumda","pas","pasda"} else "Miqdor-daraja ravishi"
                    acats = {
                        "Ravishning ma'noviy guruhlari": mn,
                        "Tuzilishi": "qo'shma",
                        "Yasalishi": "qo'shma yasalishi",
                        "Kelishik":  "Bosh kelishik",
                        "Son":       "Birlik",
                        "Egalik":    "Egalik yo'q",
                    }
                    out.append({
                        "token": phrase, "stem": pn,
                        "pos": "RR", "pos_uz": "Ravish",
                        "subtype": mn,
                        "confidence": 0.92,
                        "rule": "birikma_bir", "index": t.get("index", i),
                        "cats": acats,
                    })
                    i += 2
                    continue

            out.append(t)
            i += 1

        # indekslarni qayta raqamlash
        for k, t in enumerate(out):
            t["index"] = k
        return out


# ── Global instances ──
engine = UzbekRuleEngine()
db     = DatabaseLookup()
tagger = UzbekPOSTagger(engine, db)


# ═══════════════════════════════════════════════════════
# GROQ — stat_model so'zlarini qayta teglash
# ═══════════════════════════════════════════════════════
def groq_fill_unknowns(tokens: list) -> list:
    """DB da topilmagan (stat_model) so'zlarni Groq bilan to'g'ri teglash."""
    if not groq_client:
        return tokens

    idxs = [i for i, t in enumerate(tokens) if t.get("rule") == "stat_model"]
    if not idxs:
        return tokens

    words = [tokens[i]["token"] for i in idxs]
    prompt = (
        "O'zbek tilida quyidagi so'zlarni morfologik teglang.\n"
        "POS belgilar (datasetdagi XPOS): P=olmosh, RR=ravish, JJ=sifat, NUM=son, N=ot, V=fe'l.\n"
        "Faqat JSON array qaytaring (boshqa hech narsa yozma):\n"
        '[{"token":"...","pos":"N|V|JJ|RR|NUM|P","stem":"...","subtype":"...","confidence":0.0-1.0}]\n'
        "So'zlar: " + json.dumps(words, ensure_ascii=False)
    )
    try:
        resp = groq_client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[
                {"role": "system",
                 "content": "Sen O'zbek tili morfologiyasi ekspertisan. Faqat JSON array qaytargin."},
                {"role": "user", "content": prompt},
            ],
            temperature=0.1,
            max_tokens=800,
        )
        raw = resp.choices[0].message.content.strip()
        m = re.search(r"\[.*\]", raw, re.DOTALL)
        if m:
            parsed = json.loads(m.group())
            POS_FIX = {"ADJ":"JJ","ADV":"RR","Adj":"JJ","Num":"NUM","VB":"V"}
            for idx, res in zip(idxs, parsed):
                if isinstance(res, dict):
                    pos = str(res.get("pos", "N"))
                    pos = POS_FIX.get(pos, pos)
                    tokens[idx].update({
                        "pos":        pos,
                        "pos_uz":     POS_UZ.get(pos, "Noma'lum"),
                        "stem":       str(res.get("stem", tokens[idx]["stem"])),
                        "subtype":    str(res.get("subtype", "")),
                        "confidence": float(res.get("confidence", 0.75)),
                        "rule":       "groq_ai",
                    })
    except Exception as e:
        log.warning("Groq fill xatosi: " + str(e))

    return tokens


# ═══════════════════════════════════════════════════════
# FASTAPI
# ═══════════════════════════════════════════════════════
app = FastAPI(title="Morphological POS tagging", version="3.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_methods=["*"], allow_headers=["*"])


class TagRequest(BaseModel):
    text: str

class FilterRequest(BaseModel):
    tokens: List[dict]
    pos_types: List[str] = ["P", "RR", "JJ", "NUM"]

class AIRequest(BaseModel):
    text:     str
    tokens:   List[dict] = []
    question: str = ""

class ExportRequest(BaseModel):
    tokens:   List[dict]
    filename: str = "pos_natijalar"


@app.get("/")
async def root():
    return FileResponse(BASE_DIR / "index.html")

@app.get("/style.css")
async def style():
    return FileResponse(BASE_DIR / "style.css")

@app.get("/script.js")
async def script():
    return FileResponse(BASE_DIR / "script.js")


@app.get("/health")
async def health():
    return {
        "status":    "ok",
        "db_words":  len(db.db),
        "stat_sufs": len(db.stat.suf_cnt),
        "groq":      groq_client is not None,
        "version":   "3.0 Rule+DB+Stat+Groq",
    }


@app.post("/api/tag")
async def api_tag(req: TagRequest):
    if not req.text.strip():
        raise HTTPException(400, "Matn bo'sh")
    tokens = tagger.tag_sentence(req.text)
    tokens = groq_fill_unknowns(tokens)   # DB da yo'q so'zlarni Groq teglaydi
    stats: Dict[str, int] = {}
    for t in tokens:
        stats[t["pos"]] = stats.get(t["pos"], 0) + 1
    return {"text": req.text, "tokens": tokens, "stats": stats, "total": len(tokens)}


@app.post("/api/filter")
async def api_filter(req: FilterRequest):
    filtered = [t for t in req.tokens if t.get("pos") in req.pos_types]
    return {"tokens": filtered, "total": len(filtered)}


@app.get("/api/rules")
async def api_rules():
    e = engine
    return {"rules": {
        "olmosh": {
            "tag": "P",
            "qoida": "IF token ∈ olmoshlar → P",
            "kishilik":       sorted(e.KISHILIK),
            "korsatish":      sorted(e.KORSATISH),
            "soroq":          sorted(e.SOROQ),
            "belgilash":      sorted(e.BELGILASH),
            "bolishsizlik":   sorted(e.BOLISHSIZLIK),
            "ozlik":          sorted(e.OZLIK),
            "gumon":          sorted(e.GUMON),
            "kelishik_suf":   e.PRON_SUF[:10],
        },
        "ravish": {
            "tag": "ADV",
            "qoida": "IF token ∈ ravishlar → ADV\nIF token.endswith(suf) → ADV",
            "holat":   sorted(e.HOLAT_R),
            "payt":    sorted(e.PAYT_R),
            "orin":    sorted(e.ORIN_R),
            "miqdor":  sorted(e.MIQDOR_R),
            "maqsad":  sorted(e.MAQSAD_R),
            "sufikslar": e.ADV_SUF,
        },
        "sifat": {
            "tag": "ADJ",
            "qoida": "IF token ∈ sifatlar → ADJ\nIF token+(-roq/-gi/-dagi/-li) → ADJ",
            "rang_tus":  sorted(e.RANG_TUS),
            "maza_tam":  sorted(e.MAZA_TAM),
            "hajm":      sorted(e.HAJM),
            "xususiyat": sorted(e.XUSUSIYAT),
            "orttirma":  sorted(e.ORTTIRMA),
            "ozaytirma": sorted(e.OZAYTIRMA),
            "sufikslar": e.ADJ_SUF,
        },
        "son": {
            "tag": "NUM",
            "qoida": "IF raqam yoki son so'zi → NUM\nIF son+sufiks → NUM",
            "asosiy":   sorted(e.BASIC_NUM),
            "qoshimcha": sorted(e.NUM_EXTRA),
            "hisob_soz": sorted(e.HISOB),
            "sufiks_turlari": {k: v for k, v in e.NUM_TYPES.items()},
        },
    }}


@app.post("/api/ai")
async def api_ai(req: AIRequest):
    if not groq_client:
        raise HTTPException(503, "Groq AI ulangan emas. Server muhitida GROQ_API_KEY o'rnating.")

    # Har bir token uchun batafsil tahlil tuzamiz
    lines = []
    compounds = []
    for t in req.tokens:
        pos = t.get("pos", "?")
        if pos == "PUNCT":
            continue
        tok   = t.get("token", "")
        stem  = t.get("stem", "")
        sub   = t.get("subtype", "")
        # Qo'shma (birikma) so'zlarni aniqlaymiz
        rule  = t.get("rule", "")
        is_compound = rule.startswith("birikma") or " " in tok
        info_parts = [f"{pos}", f"lemma={stem}"]
        if sub:
            info_parts.append(f"tur={sub}")
        # Kategoriyalar (rule cats yoki DB fieldlari)
        cats = t.get("cats") or t.get("db") or {}
        for k, v in list(cats.items())[:6]:
            if v and str(v) not in ("—","∅",""):
                info_parts.append(f"{k}={v}")
        tag_line = '"' + tok + '" (' + ", ".join(info_parts) + ")"
        if is_compound:
            compounds.append(tok)
            tag_line = "[QO'SHMA] " + tag_line
        lines.append("- " + tag_line)

    analysis = "\n".join(lines) if lines else "(hali teglangan so'z yo'q)"
    comp_note = ""
    if compounds:
        comp_note = "\nMuhim: \"" + '", "'.join(compounds) + "\" — bu qo'shma (birikma) so'zlar, ulardagi har bir bo'lakni alohida emas, bitta so'z sifatida tahlil qiling.\n"

    question = req.question or "Bu gapni morfologik jihatdan to'liq tahlil qiling. Har bir so'zning POS, turi, ma'noviy guruhi, tuzilishi, yasalishini ayting."

    prompt = (
        "GAP: \"" + req.text + "\"\n\n"
        "DASTUR TAHLILI (rule + dataset):\n" + analysis + "\n"
        + comp_note +
        "\nBelgilar: P=olmosh, RR=ravish, JJ=sifat, NUM=son, N=ot, V=fe'l.\n\n"
        "SAVOL: " + question + "\n\n"
        "Qoidalar:\n"
        "1) Yuqoridagi tahlilga tayaning — undan chekinmang. Agar tahlilda qo'shma so'z bo'lsa, AI ham uni qo'shma sifatida tushuntirsin.\n"
        "2) Har bir muhim so'z/birikma uchun: XPOS, ma'noviy guruhi, tuzilishi (sodda/qo'shma/juft), yasalishi (tub/yasama)ni yozing.\n"
        "3) Olmosh bo'lsa — gapda qaysi vazifada (ot/sifat/ravish/son o'rnida) kelganini ayting.\n"
        "4) Fe'l bo'lsa — zamon, mayl, shaxs-sonni ko'rsating.\n"
        "5) Javob o'zbek tilida, qisqa va ro'yxat shaklida bo'lsin."
    )

    try:
        resp = groq_client.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[
                {"role": "system", "content":
                 "Siz o'zbek tili morfologiyasi bo'yicha ekspertsiz. "
                 "Berilgan dastur tahlilini tasdiqlab, izohlab bering; unga qarshi chiqmang. "
                 "Qo'shma (birikma) so'zlarni bitta lingvistik birlik sifatida qaraysiz."},
                {"role": "user", "content": prompt},
            ],
            temperature=0.2,
            max_tokens=1200,
        )
        return {"answer": resp.choices[0].message.content, "model": "llama-3.1-8b-instant"}
    except Exception as e:
        raise HTTPException(500, "Groq xatosi: " + str(e))


@app.post("/api/export")
async def api_export(req: ExportRequest):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl o'rnatilmagan")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "POS Natijalar"

    headers = ["#", "Token", "Lemma", "XPOS", "Turkum", "Tur / Daraja"]
    hfill = PatternFill("solid", fgColor="1D4ED8")
    hfont = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    color_map = {
        "P":  "D1FAE5", "RR":  "DBEAFE",
        "JJ": "FDE68A", "NUM": "FAE8FF",
        "N":  "F1F5F9", "V":   "FEF3C7",
    }
    for row_idx, t in enumerate(req.tokens, 2):
        fgcolor = color_map.get(t.get("pos", ""), "F9FAFB")
        fill    = PatternFill("solid", fgColor=fgcolor)
        vals = [
            row_idx - 1,
            t.get("token", ""),
            t.get("stem", ""),
            t.get("pos", ""),
            t.get("pos_uz", ""),
            t.get("subtype", ""),
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.fill = fill

    for col, w in enumerate([5, 20, 20, 8, 12, 30], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = re.sub(r"[^\w\-]", "_", req.filename) + ".xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=" + fname},
    )


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    groq_status = "ulandi" if groq_client else "yo'q (GROQ_API_KEY o'rnating)"
    log.info("=" * 58)
    log.info("Morphological POS tagging v3.0")
    log.info("DB: " + str(len(db.db)) + " so'z | Stat: " + str(len(db.stat.suf_cnt)) + " naqsh | Groq: " + groq_status)
    log.info("Server: http://0.0.0.0:" + str(port))
    log.info("=" * 58)
    uvicorn.run("server:app", host="0.0.0.0", port=port, reload=False)
